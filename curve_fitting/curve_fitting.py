import argparse
import collections
import glob
import os
import re
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import sklearn.metrics
import torch
from matplotlib.widgets import CheckButtons
from openpyxl.reader.excel import load_workbook
from scipy.optimize import curve_fit, least_squares
# from numdifftools import Jacobian
from scipy.stats import chisquare
from sklearn.metrics import mean_absolute_error, mean_squared_error
from torch.autograd.functional import jacobian
from torch import tensor
from scipy.signal import convolve


class CurveFitting(object):

    def __init__(self, init_params, bounds=(0, np.inf), use_jac=False, verbose=False):
        self.init_params = init_params
        self.bounds = bounds
        self.use_jac = use_jac
        self.verbose = verbose
        self.fitted_params = init_params

    @staticmethod
    def forward_model(x, p):

        y_comps = []
        if len(p) % 2:
            # print('p[0]', p, p[0])
            y_comps.append((np.ones((x.shape[0])) * p[0]))
            p = p[1:]

        for comp, (tau, beta) in enumerate(zip(p[::2], p[1::2])):
            # print('tau, beta', tau, beta)
            y_comps.append(beta * np.exp(-x/tau))

        # print('comps shape', np.asarray(comps).shape) (n_time_params + DC, n_points)
        # print(sum(comps).shape) # (n_points, )

        return np.sum(np.array(y_comps, dtype=np.float32), axis=0), \
            np.asarray(y_comps, dtype=np.float32)

    @staticmethod
    def _backward_model(p, x, y):

        grads = []
        if len(p) % 2:
            grads.append(np.ones(x.shape[0]))
            p = p[1:]

        for tau, beta in zip(p[::2], p[1::2]):
            # print('tau, beta', tau, beta)
            grads.append((beta * np.exp(-x / tau) / -tau))
            grads.append((np.exp(-x / tau)))

        # print('grads shape jac', np.array(grads).T.shape)
        return np.array(grads).T

    @staticmethod
    def __lsq_func(p, x, y):

        y_fitted, _ = CurveFitting.forward_model(x, p)

        # res = (1 / (x.shape[0] - len(p))) * np.sum(np.square(y_fitted - y)/y_fitted)
        # res = (1 / (x.shape[0] - len(p) + 4)) * np.square(y_fitted - y)/y_fitted
        # res = (y - y_fitted) / np.sqrt(y)
        res = np.divide((y - y_fitted), np.sqrt(y))
        # print('res shape: ', res.shape)

        return res

    def fit_lsq(self, x, y):
        # bounds = ([2.5, 0, 0, 0, 0, 0, 0], [np.inf, np.inf, np.inf, np.inf, np.inf, np.inf, np.inf])
        # bounds = ([0, 0, 0, 0, 14, 0], [np.inf, np.inf, np.inf, np.inf, 19, np.inf])
        # return least_squares(self._func_lsq, x0=self.p, jac=self._d_func_lsq, args=(x, y))
        # return least_squares(self._func_lsq, x0=self.p, jac=self._d_func_lsq_torch, args=(x, y))
        # return least_squares(self._func_lsq, x0=self.p, jac=self._d_func_lsq, args=(x, y), method='lm')
        # return least_squares(self._func_lsq, x0=self.p, args=(x, y), method='lm')
        # res_lsq = least_squares(self.__lsq_func, x0=self.__init_params, args=(x, y)) # best
        # return least_squares(self._func_lsq, x0=self.p, bounds=(0, np.inf), args=(x, y))  # best
        if self.use_jac is True:
            res_lsq = least_squares(self.__lsq_func, x0=self.init_params, jac=self._backward_model,
                                    bounds=self.bounds, args=(x, y))
        else:
            res_lsq = least_squares(self.__lsq_func, x0=self.init_params, bounds=self.bounds, args=(x, y))
        # return least_squares(self._func_lsq, x0=self.p, bounds=(0, np.inf), args=(x, y), loss='arctan') # trial
        # return least_squares(self._func_lsq, x0=self.p, jac='cs', bounds=(0, np.inf), args=(x, y)) # best
        # return least_squares(self._func_lsq, x0=self.p, args=(x, y), jac=self._d_func_lsq, method='lm')
        if self.verbose is True:
            print('res_lsq', res_lsq)

        # self.fitted_params = res_lsq['x']
        return res_lsq['x']

    @staticmethod
    def red_chi_sqr(y, y_fitted, p):
        n_params = p.shape[0]
        return (1 / (y.shape[0] - n_params + (n_params // 2))) * np.sum(np.square(y-y_fitted)/y)

    @staticmethod
    def rel_contrib_int(p):

        n_comp = len(p) // 2
        if len(p) % 2:
            p = p[1:]
        p = np.array(p, dtype=np.float32)

        return (100 * p[::2][n_comp-1] * p[1::2][n_comp-1]) / np.sum((p[::2] * p[1::2]))

    @staticmethod
    def rel_contrib_conc(p):

        n_comp = len(p) // 2
        if len(p) % 2:
            p = p[1:]
        p = np.array(p, dtype=np.float32)

        return 100 * p[1::2][n_comp-1] / np.sum(p[1::2])

    @staticmethod
    def matrices(y, y_fitted, fitted_params) -> tuple:
        ':returns rel_contrib_int, rel_contrib_conc, red_chi_sqr'
        return CurveFitting.rel_contrib_int(fitted_params), \
            CurveFitting.rel_contrib_conc(fitted_params), \
            CurveFitting.red_chi_sqr(y, y_fitted, fitted_params)


def plot_curves(x, y=None, y_fitted=None, y_comps=None):

    # Create the figure and subplot
    fig, ax = plt.subplots()

    # max_idx = np.argmax(y)
    # print('max idx', max_idx, np.max(y), y[max_idx], y[max_idx + 1])
    # ax.plot(x, y, label='original')
    # ax.plot(x, y_fitted, label='yfitted')
    # for idx, comp in enumerate(y_comps):
    #     # print('comp: ', len(comp))
    #     ax.plot(x, comp, label='comp' + str(idx))


    # Plot the curves
    curves = []
    if y is not None:
        curves.append(ax.plot(x, y, label='org')[0])
    if y_fitted is not None:
        curves.append(ax.plot(x, y_fitted, label='yfitted')[0])
    for idx, comp in enumerate(y_comps):
        # print('comp: ', len(comp))
        curves.append(ax.plot(x, comp, label='comp' + str(idx))[0])

    # Create the checkbox widget
    rax = plt.axes([0.7, 0.7, 0.1, 0.15])
    labels = [str(line.get_label()) for line in curves]
    check = CheckButtons(rax, labels, [True] * len(labels))

    # Define the function to update the plot when the checkboxes are clicked
    def update(val):
        for i, line in enumerate(curves):
            if check.get_status()[i]:
                line.set_visible(True)
            else:
                line.set_visible(False)
        plt.draw()

    # Connect the checkbox widget to the update function
    check.on_clicked(update)

    # Show the plot
    plt.legend()
    plt.show()


def preprocess(DATA, tail_trim=400, zero_trim=True) -> dict:

    # DATA = np.array(list(DATA.values()), dtype=np.float32)
    # print('DATA', DATA.shape)  # (n_samples, points, (x, y))

    x_offset = 1
    y_offset = 1

    # PROC_DATA = []
    for idx, (filename, data) in enumerate(DATA.items()):

        y = data[:, 1]
        x = data[:, 0]

        # plt.plot(x, y)
        # plt.show()

        if zero_trim:
            # print('filename in preprocess: ', filename)
            # Remove zero values
            zero_ind = np.where(y == 0)
            # print('zero_ind', zero_ind)
            # x, y -> (n_points - n_zero_points)
            y = np.delete(y, zero_ind)
            x = data[:y.shape[0], 0]
            # print(' x y shape', x.shape, y.shape) # n_samples - zero values

            # trim head (i.e., start from max) and tail
            max_idx = np.argmax(y)
            # print('max idx', max_idx, np.max(y), y[max_idx], y[max_idx + 1])
            x = x[x_offset:x_offset + tail_trim]
            y = y[max_idx + y_offset:max_idx + y_offset + tail_trim]

        else:

            # trim head (i.e., start from max) and tail
            max_idx = np.argmax(y)
            # print('max idx', max_idx, np.max(y), y[max_idx], y[max_idx + 1])
            x = x[x_offset:x_offset + tail_trim]
            y = y[max_idx + y_offset:max_idx + y_offset + tail_trim]

            zero_ind = np.where(y == 0)
            # print('zero_ind', zero_ind, len(zero_ind[0]))
            # # x, y -> (n_points - n_zero_points)
            # y = np.delete(y, zero_ind)
            if not len(zero_ind[0]) == 0:
                # print('zero_ind inside', zero_ind[0])
                y = y[:zero_ind[0][0]]
                x = x[:y.shape[0]]
                # exit()

        # else:
        #     y = data[:, 1]
        #     x = data[:, 0]
        #
        # plt.plot(x, y)
        # plt.show()


        # plt.scatter(x, y)
        # plt.show()

        DATA[filename] = np.asarray([x, y], dtype=np.float32)

        # print('x, y shape', x.shape, y.shape)
        # plt.plot(DATA[filename][0], DATA[filename][1])
        # plt.show()

    return DATA


def file_parser(file):

    file_data = open(file, 'r')
    lines = file_data.readlines()

    data_flag = False
    data = []
    for line in lines:

        if data_flag is False:
            # print('line, length', len(line))
            if len(line) == 1:
                data_flag = True
        else:
            # print('line: ', line.split(',')[0])
            data.append((line.split(',')[0], line.split(',')[1]))

    return data


def get_data(_path, pattern):

    DATA = {}

    if os.path.isdir(_path):

        for pat in pattern:

            for filename in Path(_path).rglob(pat):
            # for filename in sorted(glob.iglob((_path + pat), recursive=True)):
            # for filename in sorted(glob.glob((_path + pat))):
            # for filename in sorted(glob.glob((_PATH + '/*_Decay*.txt'))):
            # for filename in sorted(glob.glob((_PATH + '/*_Decay.txt'))):

                # if not FILE_NAME in file:
                #     continue
                # print('filename: ', filename)

                data = file_parser(filename)
                data = np.array(data, dtype=np.float32)
                # data[:, 1] = convolve(irf[:, 1], data[:, 1], mode='same')
                # print('data shape:', data.shape)
                DATA[os.path.basename(filename).strip('.txt')] = data

        return DATA

    elif os.path.isfile(_path):

        DATA[os.path.basename(_path).strip('.txt')] = np.array(file_parser(_path), dtype=np.float32)
        return DATA

    else:
        raise FileNotFoundError


def post_process(PARAMS):

    df = pd.DataFrame.from_dict(PARAMS, orient='index')
    df.columns = ['alpha', 'tau1', 'beta1', 'tau2', 'beta2', 'tau3', 'beta3',
                  'rel_fl_int', 'rel_concentration', 'red_chi_squares']

    RED_CHI_SQUARES = df.iloc[:, -1].to_numpy()
    mae_red_chi_square = mean_absolute_error(np.ones(RED_CHI_SQUARES.shape),
                                             RED_CHI_SQUARES)
    mse_red_chi_square = mean_squared_error(np.ones(RED_CHI_SQUARES.shape),
                                            RED_CHI_SQUARES)

    # mean_values = df.iloc[:, :-1].mean()
    mean_values = df.mean()
    std_dev = df.std()
    df.loc['mean'] = mean_values
    df.loc['std'] = std_dev
    df._set_value('mean', 'red_chi_squares', mae_red_chi_square)
    df._set_value('std', 'red_chi_squares', mse_red_chi_square)

    # print('average_params:\n', mean_values[:-1])
    print('min values:\n', df.min(axis=0))
    print('max values:\n', df.max(axis=0))
    print('average:\n', df.loc['mean'])
    print('std:\n', df.loc['std'])
    # print('chisquares MEA: {:.2f}'.format(mae_red_chi_square))
    print('n_dodgy chi squares: {:.4f}'.format(len(np.where(np.logical_or(RED_CHI_SQUARES >= 1.2, RED_CHI_SQUARES <= 0.85))[0])/len(RED_CHI_SQUARES)))
    df = df.round(4)
    # print('summary:\n', df)

    return df


def write_to_sheet(df, wb_name, sheet_name):

    if os.path.isfile(wb_name):
        # book = load_workbook(SPREADSHEET)
        # writer = pd.ExcelWriter(wb_name, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        writer = pd.ExcelWriter(wb_name, engine='openpyxl', mode='a')
        writer.workbook = openpyxl.load_workbook(wb_name)
        df.to_excel(writer, sheet_name=sheet_name)
        writer.close()

    else:
        writer = pd.ExcelWriter(wb_name, engine='openpyxl')
        df.to_excel(writer, sheet_name=sheet_name)
        writer.close()


def main(args):

    DATA = get_data(args.path, args.pattern)

    if not bool(DATA):
        raise KeyError('DATA is empty')

    DATA = preprocess(DATA, tail_trim=args.tail_trim, zero_trim=args.zero_trim)
    print('n_files parsed', len(DATA.keys()))
    # exit()
    DATA = collections.OrderedDict(sorted(DATA.items(), key=lambda _i: _i[0].lower()))

    curve_fitting = CurveFitting(args.init_params, args.bounds,
                                 use_jac=args.use_jac)
    PARAMS = {}
    for i, (filename, [x, y]) in enumerate(DATA.items()):

        fitted_params = curve_fitting.fit_lsq(x, y)

        y_fitted, y_comps = CurveFitting.forward_model(x, fitted_params)

        matrices = curve_fitting.matrices(y, y_fitted, fitted_params)
        print('filename n_points: ', filename, x.shape)
        print('fitted_params', fitted_params)
        print('rel_fl_int {:.2f}, rel_concentration, {:.2f}, red_chi_sqr {:.2f}\n'.format(*matrices))

        if args.plot:
            plot_curves(x, y=y, y_fitted=y_fitted, y_comps=y_comps)
        # exit()

        PARAMS[filename] = [*fitted_params,
                            *matrices]

    df = post_process(PARAMS)

    if args.write:
        write_to_sheet(df, args.wb_name, args.sheet_name)


if __name__ == '__main__':

    # PATH = '/home/bappadityadebnath/Documents/Curve_fitting'
    # PATH = '/home/bappadityadebnath/Downloads/PPEye001/Fluorimeter Data/13102022'
    # PATH = '/home/bappadityadebnath/datasets/matt_data/PPEye002'
    # PATH = '/home/bappadityadebnath/datasets/matt_data/PPEye005/Fluorimeter/'
    # PATH = '/home/bappadityadebnath/datasets/matt_data/PPEye003/PPEye003/Fluorimeter/'
    # print('PATH', PATH)
    # PATH = '/home/bappadityadebnath/datasets/matt_data/PPEye_Stage_1_data/'
    # IRF_FILE = '/home/bappadityadebnath/datasets/matt_data/IRF/IRF_BW10_500ns.txt'
    # PATH = '/home/bappadityadebnath/datasets/matt_data/PPEye_Stage_1_data/PPEye001/Fluorimeter'
    # PATH = '/home/bappadityadebnath/datasets/matt_data/PPEye_Stage_1_data/PPEye002/Fluorimeter/'
    # PATH = '/home/bappadityadebnath/datasets/matt_data/PPEye_Stage_1_data/PPEye003/Fluorimeter'
    # PATH = '/home/bappadityadebnath/datasets/matt_data/PPEye_Stage_1_data/PPEye004/Fluorimeter'
    # PATH = '/home/bappadityadebnath/datasets/matt_data/PPEye_Stage_1_data/PPEye005/Fluorimeter'
    # PATH = '/home/bappadityadebnath/datasets/matt_data/PPEye_Stage_1_data/PPEye006/Fluorimeter'
    # PATH = '/home/bappadityadebnath/datasets/matt_data/PPEye_Stage_1_data/PPEye007/Fluorimeter'
    # PATH = '/home/bappadityadebnath/datasets/matt_data/PPEye_Stage_1_data/PPEye008/Fluorimeter'
    # PATH = '/home/bappadityadebnath/datasets/matt_data/PPEye_Stage_1_data/PPEye009/Fluorimeter'
    # PATH = '/home/bappadityadebnath/datasets/matt_data/PPEye_Stage_1_data/PPEye010/Fluorimeter'
    # PATH = '/home/bappadityadebnath/datasets/matt_data/PPEye_Stage_1_data/PPEye011/Fluorimeter'
    # PATH = '/home/bappadityadebnath/datasets/matt_data/PPEye_Stage_1_data/PPEye012/Fluorimeter'
    PATH = '/home/bappadityadebnath/datasets/matt_data/PPEye_Stage_1_data/PPEye013/Fluorimeter'

    # PATH = '/home/bappadityadebnath/datasets/matt_data/PPEye_Stage_1_data_control/PPEyeControl001/Fluorimeter'
    # PATH = '/home/bappadityadebnath/datasets/matt_data/PPEye_Stage_1_data_control/PPEyeControl002/Fluorimeter'
    # PATH = '/home/bappadityadebnath/datasets/matt_data/PPEye_Stage_1_data_control/PPEyeControl003/Fluorimeter'
    # PATH = '/home/bappadityadebnath/datasets/matt_data/PPEye_Stage_1_data_control/PPEyeControl004/Fluorimeter'

    # PATH = '/home/bappadityadebnath/datasets/matt_data/PPEye_Stage_1_data/PPEye006/Fluorimeter/PPEye006-002-11-Decay625.txt'
    # PATH = '/home/bappadityadebnath/datasets/matt_data/PPEye_Stage_1_data/PPEye006/Fluorimeter/PPEye006-002-11-Decay635.txt'
    # PATH = '/home/bappadityadebnath/datasets/matt_data/PPEye_Stage_1_data/PPEye006/Fluorimeter/PPEye006-002-9-Decay635.txt'

    sheet_name = 'PPEye_Stage_1_data'
    # sheet_name = re.search(r"/(PPEye(Control)?0[0-9][0-9])/", PATH)[1]
    # print('sheet_name', sheet_name)
    # exit()

    # pattern
    # pattern = '/*-Decay*.txt'
    # pattern = '/*_Decay*.txt'
    # pattern = '/*_Decay.txt'
    # pattern = ['/*-Decay*.txt', '/*_Decay*.txt', '/*_Decay.txt']
    pattern = ['*-Decay*.txt', '*_Decay*.txt', '*_Decay.txt']

    # in order of tau, beta
    # p = [17, 1000]
    # p = [1.5, 0.01, 14, 0.1]
    # init_params = [10, 5, 5000, 10, 1000]
    # init_params = [0, .7, 6700, 3.5, 2670, 17, 1150]  # previous
    # init_params = [10, .7, 6700, 3.5, 2670, 17, 1150]  # best
    # init_params = [3, .6, 5791, 2.5, 1110, 15, 220]  # current trial
    # init_params = [3, .6, 5791, 3.5, 1110, 15, 220]  # best 22_03_23 11:26
    init_params = [0, .65, 12000, 3.65, 1370, 16, 2046]  # current trial
    # init_params = [.6, 5791, 2.5, 1110, 15, 220]  # current trial no dc
    # init_params = [10, .7, 6700, 3, 2670] # best for control
    # init_params = [10, .7, 5791, 3.0, 1110, 5, 1000, 17, 220]
    # init_params = [0, .1, 5000, 1, 2500, 10, 1000]
    # init_params = [.7, 6700, 3.5, 2670, 17, 1150]
    # init_params = [10, 0.001, 10000, .7, 6700, 3.5, 2670, 17, 1150]
    # p = [5, 1000, 5, 1000, 10, 1000]

    # bounds
    bounds = ([0, 0, 0, 0, 0, 14, 0], [np.inf, 1, np.inf, 7, np.inf, 19, np.inf])
    # bounds = ([0, 0, 0, 0, 14, 0], [1, np.inf, 7, np.inf, 19, np.inf]) # no dc
    # bounds = ([0, 0, 0, 0, 0, 14, 0], [np.inf, 1, np.inf, 7, np.inf, 19, np.inf]) #best 22/03/23
    # bounds = ([2.5, 0, 0, 0, 0, 14, 0], [np.inf, 1, np.inf, 7, np.inf, 19, np.inf])  # trial
    # bounds = (0, np.inf)
    # bounds = ([0, 0, 0, 0, 0, 14, 0], [np.inf, np.inf, np.inf, np.inf, np.inf, 19, np.inf])

    parser = argparse.ArgumentParser()

    parser.add_argument("--path", default=PATH)
    parser.add_argument("--pattern", default=pattern)
    parser.add_argument("--tail_trim", default=400)
    parser.add_argument("--zero_trim", default=False)
    parser.add_argument("--init_params", default=init_params)
    parser.add_argument("--bounds", default=bounds)
    parser.add_argument("--use_jac", default=False)
    parser.add_argument("--verbose", default=True)
    parser.add_argument("--plot", default=False)
    parser.add_argument("--write", default=False)
    parser.add_argument("--wb_name", default="./workbooks/PPEye_stage_1_till_zero.xlsx")
    parser.add_argument("--sheet_name", default=sheet_name)

    args = parser.parse_args()
    # print('args', args)

    main(args)

